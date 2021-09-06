import openpyxl
import os
class ExcelObject():
    def __init__(self, path=None, file="excel_data.xlsx", read_only=True):
        if not path:
            path=os.path.join(os.path.split(os.path.dirname(__file__))[0],'excel_data')
        if read_only:
            self.wb=openpyxl.load_workbook(os.path.join(path, file), read_only=True)
        else:
            self.wb=openpyxl.load_workbook(os.path.join(path, file))


    def read(self, table, rows=None):
        self.tables_name=self.wb.sheetnames  
        if type(table)==str:
            self.table=self.wb[table]
        elif type(table)==int:     
            self.table=self.wb[self.tables_name[table]]
        else:
            raise ValueError("ExcelObject.read()必须指定字符串表名，或者整数表序号")
        self.rows=[[x.value for x in y ] for y in self.table.rows]
        if not rows:
            self.data = self.rows
        elif type(rows)==list:
            self.data=[]
            if type(rows[0])==int:
                for i in rows:
                    self.data.append(self.rows[i])
            elif type(rows[0])==list:
                for i in rows:
                    line=[]
                    for j in range(i[0],i[1]+1):
                        line.append(self.rows[j])
                    self.data.append(line)
            else:
                raise ValueError("ExcelObject.rows必需传入[int,int...]或者[[int,int],[int,int]..]")
        else:
            raise ValueError("ExcelObject.rows必需传入[int,int...]或者[[int,int],[int,int]..]")
        return self.data

    def get_tables_name(self):
        return self.wb.sheetnames  

    def get_table_example(self, table, label=0):
        return self.read(table, rows=[label])[0]

#c=ExcelObject(file="souhu_login.xlsx")
#print(c.read("login"))
#print(c.get_table_example("测试"))
#print(os.path.join(os.path.split(os.path.dirname(__file__))[0],'excel_data'))
#raise ValueError("ExcelObject.read_table()必须制定表名，或者表序号")
